"""
fmeaParser.py
─────────────────────────────────────────────────────────────────────────────
FMEA spreadsheet (CSV / Excel) → canonical FMEA record list.

Each non-empty input row becomes a dict with normalised canonical field names.
Column headings are matched against a configurable regex map so that plant-
specific naming variations (e.g. "SEV" vs "Severity") are handled
transparently.

Supported formats
─────────────────
• CSV  — any delimiter recognised by csv.Sniffer; UTF-8 or latin-1
• .xlsx — via openpyxl (already a project dependency)
• .xls  — via xlrd  (already a project dependency)
• Multi-sheet workbooks — all sheets parsed; records carry ``_sheet``

Output schema (canonical keys)
───────────────────────────────
Required (hard-fail if absent):
  fmea_source_ref     str   source filename (set from the input path)
  component_type      str   equipment class (e.g. "centrifugal_pump")
  failure_mode_id     str   FM:<slug(component_type)>:<slug(failure_mode_name)>
  failure_mode_name   str   human-readable failure mode label

Optional:
  failure_mechanism        str    physical mechanism (fatigue, corrosion, …)
  local_effect             str    local / end effect description
  severity                 int    1–10
  occurrence               int    1–10
  detection                int    1–10
  rpn                      int    explicit or derived = S × O × D
  expected_latency_min_hours  float  converted from min_days × 24
  expected_latency_max_hours  float  converted from max_days × 24
  expected_anomaly_pattern    str    normalised to enum values
  expected_symptoms           list[str]  split from local_effect text
  corrective_actions          list[str]
  notes                       str
  _sheet                      str    Excel sheet name; None for CSV
  _row_index                  int    1-based row number after header
"""

from __future__ import annotations

import csv
import logging
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    from doc_parsers.fmea_normalizer import normalize_fmea_records
except ModuleNotFoundError:
    from dackar.RCA.doc_parsers.fmea_normalizer import normalize_fmea_records  # type: ignore

LOGGER = logging.getLogger("fmeaParser")
if not LOGGER.handlers:
    _ch = logging.StreamHandler()
    _ch.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(name)s | %(message)s"))
    LOGGER.addHandler(_ch)
LOGGER.setLevel(logging.INFO)


# ---------------------------------------------------------------------------
# Default column name map
# Keys are canonical field names; values are lists of regex patterns matched
# case-insensitively against the stripped column header.
# Plants override by passing ``column_map_override`` to ``parse_fmea_file``.
# ---------------------------------------------------------------------------

DEFAULT_COLUMN_MAP: Dict[str, List[str]] = {
    "component_type": [
        r"component[\s_-]?type",
        r"equipment[\s_-]?type",
        r"item[\s_-]?type",
        r"^item$",
        r"^component$",
        r"^equipment$",
        r"asset[\s_-]?type",
    ],
    "failure_mode_name": [
        r"failure[\s_-]?mode",
        r"potential[\s_-]?failure[\s_-]?mode",
        r"mode[\s_-]?of[\s_-]?failure",
        r"^fm$",
        r"^fm[\s_-]?name$",
    ],
    "item_function": [
        r"item[\s_-]?function",
        r"function",
        r"system[\s_-]?element",
    ],
    "failure_mechanism": [
        r"failure[\s_-]?mechanism",
        r"mechanism",
        r"physical[\s_-]?mechanism",
        r"cause[\s_-]?mechanism",
        r"degradation[\s_-]?mechanism",
    ],
    "local_effect": [
        r"local[\s_-]?effect",
        r"failure[\s_-]?effect",
        r"effect",
        r"symptom",
        r"consequence",
    ],
    "system_effect": [
        r"system[\s_-]?effect",
        r"sub[\s_-]?system[\s_-]?effect",
        r"next[\s_-]?higher[\s_-]?effect",
        r"higher[\s_-]?effect",
    ],
    "end_effect": [
        r"end[\s_-]?effect",
        r"mission[\s_-]?effect",
        r"safety[\s_-]?effect",
        r"effect[\s_-]?on[\s_-]?plant",
    ],
    "potential_causes": [
        r"potential[\s_-]?causes?",
        r"cause\(s\)",
        r"root[\s_-]?causes?",
        r"failure[\s_-]?causes?",
    ],
    "detection_method": [
        r"detection[\s_-]?method",
        r"current[\s_-]?controls?[\s_-]?\(detection\)",
        r"how[\s_-]?detected",
    ],
    "severity": [
        r"^severity$",
        r"^sev$",
        r"^s$",
        r"severity[\s_-]?rating",
        r"severity[\s_-]?\(s\)",
    ],
    "occurrence": [
        r"^occurrence$",
        r"^occ$",
        r"^o$",
        r"occurrence[\s_-]?rating",
        r"occurrence[\s_-]?\(o\)",
        r"frequency",
    ],
    "detection": [
        r"^detection$",
        r"^det$",
        r"^d$",
        r"detection[\s_-]?rating",
        r"detection[\s_-]?\(d\)",
    ],
    "detection_rating": [
        r"detection[\s_-]?rating",
        r"current[\s_-]?controls?[\s_-]?\(detection\)",
    ],
    "rpn": [
        r"^rpn$",
        r"risk[\s_-]?priority[\s_-]?number",
        r"risk[\s_-]?number",
    ],
    "expected_latency_min_days": [
        r"latency[\s_-]?min",
        r"min[\s_-]?latency",
        r"minimum[\s_-]?latency",
        r"progression[\s_-]?min",
        r"min[\s_-]?days",
        r"latency[\s_-]?min[\s_-]?\(days?\)",
    ],
    "expected_latency_max_days": [
        r"latency[\s_-]?max",
        r"max[\s_-]?latency",
        r"maximum[\s_-]?latency",
        r"progression[\s_-]?max",
        r"max[\s_-]?days",
        r"latency[\s_-]?max[\s_-]?\(days?\)",
    ],
    "expected_anomaly_pattern": [
        r"anomaly[\s_-]?pattern",
        r"telemetry[\s_-]?pattern",
        r"signal[\s_-]?pattern",
        r"expected[\s_-]?pattern",
    ],
    "corrective_actions": [
        r"corrective[\s_-]?action",
        r"recommended[\s_-]?action",
        r"action[\s_-]?item",
        r"mitigation",
    ],
    "safety_function_impact": [
        r"safety[\s_-]?function[\s_-]?impact",
        r"affected[\s_-]?safety[\s_-]?functions?",
        r"safety[\s_-]?impact",
    ],
    "tech_spec_applicability": [
        r"tech[\s_-]?spec[\s_-]?applicability",
        r"lco[\s_-]?applicability",
        r"technical[\s_-]?spec",
    ],
    "failure_rate": [
        r"failure[\s_-]?rate",
        r"lambda",
        r"^λ$",
    ],
    "failure_mode_ratio": [
        r"failure[\s_-]?mode[\s_-]?ratio",
        r"alpha",
        r"^α$",
    ],
    "mission_time_hours": [
        r"mission[\s_-]?time",
        r"operating[\s_-]?hours",
        r"mission[\s_-]?hours",
        r"\bt\b",
    ],
    "criticality": [
        r"criticality",
        r"criticality[\s_-]?rank",
        r"class[\s_-]?[ivx]+",
    ],
    "notes": [
        r"^notes?$",
        r"^comment",
        r"^remarks?$",
        r"additional[\s_-]?info",
    ],
    "fmea_revision_date": [
        r"fmea[\s_-]?revision[\s_-]?date",
        r"revision[\s_-]?date",
        r"last[\s_-]?updated",
    ],
}

# Profile-specific header patterns (prepended before defaults).
PROFILE_COLUMN_MAPS: Dict[str, Dict[str, List[str]]] = {
    "mil_std_1629a": {
        "component_type": [
            r"item",
            r"assembly",
            r"line[\s_-]?replaceable[\s_-]?unit",
            r"lru",
        ],
        "failure_mode_name": [
            r"potential[\s_-]?failure[\s_-]?mode",
            r"failure[\s_-]?mode[\s_-]?description",
        ],
        "failure_mechanism": [
            r"cause[\s_-]?of[\s_-]?failure",
            r"failure[\s_-]?cause",
            r"mechanism[\s_-]?of[\s_-]?failure",
        ],
        "criticality": [
            r"criticality[\s_-]?class",
            r"class[\s_-]?[ivx]+",
        ],
        "failure_rate": [
            r"failure[\s_-]?rate[\s_-]?\(λ\)",
            r"lambda[\s_-]?\(λ\)",
        ],
        "mission_time_hours": [
            r"mission[\s_-]?time[\s_-]?\(h(rs)?\)",
            r"mission[\s_-]?duration",
        ],
    },
    "aiag_5th": {
        "detection_rating": [
            r"detection[\s_-]?control[\s_-]?rating",
            r"detection[\s_-]?ranking",
        ],
        "occurrence": [
            r"occurrence[\s_-]?ranking",
        ],
    },
    "nuclear_generic": {
        "safety_function_impact": [
            r"affected[\s_-]?safety[\s_-]?function",
            r"safety[\s_-]?function",
        ],
        "tech_spec_applicability": [
            r"technical[\s_-]?spec[\s_-]?applicability",
            r"lco",
        ],
    },
}

# Allowed values for expected_anomaly_pattern (from kg_context failure_modes schema).
_ANOMALY_PATTERN_ENUM = frozenset({
    "step_change", "gradual_drift", "spike",
    "oscillation", "dropout", "sustained_exceedance", "unknown",
})

# Keyword → canonical anomaly pattern normalisation map.
_ANOMALY_PATTERN_KEYWORDS: Dict[str, str] = {
    "step": "step_change",
    "step change": "step_change",
    "drift": "gradual_drift",
    "gradual": "gradual_drift",
    "gradual drift": "gradual_drift",
    "ramp": "gradual_drift",
    "spike": "spike",
    "transient": "spike",
    "impulse": "spike",
    "oscillat": "oscillation",
    "fluctuat": "oscillation",
    "cycle": "oscillation",
    "dropout": "dropout",
    "drop out": "dropout",
    "loss of signal": "dropout",
    "signal loss": "dropout",
    "exceedance": "sustained_exceedance",
    "sustained": "sustained_exceedance",
    "high": "sustained_exceedance",
    "overload": "sustained_exceedance",
}

# Effect text delimiters used to split local_effect into symptom list.
_EFFECT_SPLIT_RE = re.compile(r"[;,/|]|\band\b|\bor\b", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _slug(text: str) -> str:
    """Return a lowercase, underscore-separated identifier-safe string."""
    text = unicodedata.normalize("NFKC", text).lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s\-]+", "_", text)
    return text.strip("_")


def _norm(value: Any) -> str:
    """Strip and lower a cell value."""
    return " ".join(str(value).split()).lower().strip() if value is not None else ""


def _to_int(value: Any) -> Optional[int]:
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def _to_float(value: Any) -> Optional[float]:
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def _resolve_anomaly_pattern(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    normed = _norm(raw)
    if normed in _ANOMALY_PATTERN_ENUM:
        return normed
    for keyword, canonical in _ANOMALY_PATTERN_KEYWORDS.items():
        if keyword in normed:
            return canonical
    return "unknown"


def _split_effect_to_symptoms(effect_text: Optional[str]) -> List[str]:
    if not effect_text:
        return []
    parts = _EFFECT_SPLIT_RE.split(effect_text)
    return [p.strip() for p in parts if p.strip()]


def _split_actions(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    parts = re.split(r"[;|]|\n|\d+\.", str(raw))
    return [p.strip() for p in parts if p.strip()]


def _split_causes(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    parts = re.split(r"[;|/]|\band\b", str(raw), flags=re.IGNORECASE)
    return [p.strip() for p in parts if p.strip()]


def _build_column_map(
    *,
    profile_name: str,
    column_map_override: Optional[Dict[str, List[str]]] = None,
) -> Dict[str, List[str]]:
    # Merge precedence: override > profile > defaults
    merged_map: Dict[str, List[str]] = {}
    profile_map = PROFILE_COLUMN_MAPS.get((profile_name or "").strip().lower(), {})
    for canonical, patterns in DEFAULT_COLUMN_MAP.items():
        extra = (column_map_override or {}).get(canonical, [])
        profile_patterns = profile_map.get(canonical, [])
        merged_map[canonical] = list(extra) + list(profile_patterns) + list(patterns)
    for canonical, patterns in profile_map.items():
        if canonical not in merged_map:
            merged_map[canonical] = list(patterns)
    for canonical, patterns in (column_map_override or {}).items():
        if canonical not in merged_map:
            merged_map[canonical] = list(patterns)
    return merged_map


# ---------------------------------------------------------------------------
# Column resolver
# ---------------------------------------------------------------------------

class FmeaColumnResolver:
    """Resolve actual spreadsheet column headers to canonical field names.

    Resolution is purely regex-based: each header is tested against every
    pattern list in the column map.  The first match wins; if a header matches
    multiple canonical fields the first canonical field (alphabetical order)
    wins and a warning is logged.

    Args:
        column_map: Merged column map (defaults + overrides).
    """

    def __init__(self, column_map: Dict[str, List[str]]):
        self._map = column_map

    def resolve(self, headers: Sequence[str]) -> Dict[str, int]:
        """Return a dict mapping canonical field name → 0-based column index.

        Unrecognised headers are silently ignored.

        Args:
            headers: Raw header strings from the spreadsheet.

        Returns:
            ``{canonical_field: col_index}`` for every resolved column.
        """
        resolved: Dict[str, int] = {}
        for col_idx, raw_header in enumerate(headers):
            header = _norm(raw_header)
            if not header:
                continue
            matched = False
            for canonical, patterns in self._map.items():
                for pat in patterns:
                    if re.fullmatch(pat, header, re.IGNORECASE):
                        if canonical in resolved:
                            LOGGER.warning(
                                "Column '%s' matches canonical field '%s' which is already "
                                "resolved to column %d — skipping duplicate.",
                                raw_header, canonical, resolved[canonical],
                            )
                        else:
                            resolved[canonical] = col_idx
                        matched = True
                        break
                if matched:
                    break
        return resolved

    def validate_required(self, resolved: Dict[str, int], source: str) -> None:
        """Raise :class:`ValueError` if required fields are missing.

        Args:
            resolved: Output of :meth:`resolve`.
            source: Human-readable source description for the error message.

        Raises:
            ValueError: If ``component_type`` or ``failure_mode_name`` cannot
                be resolved, with a list of all detected headers included.
        """
        missing = [
            f
            for f in ("component_type", "failure_mode_name", "failure_mechanism")
            if f not in resolved
        ]
        if missing:
            raise ValueError(
                f"FMEA parse error in '{source}': required column(s) {missing} could not be "
                f"resolved. Detected canonical fields: {sorted(resolved.keys())}. "
                f"Check that the spreadsheet has columns matching the DEFAULT_COLUMN_MAP "
                f"patterns or supply a column_map_override."
            )


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

def _build_record(
    cells: Dict[str, Any],
    row_index: int,
    fmea_source_ref: str,
    sheet: Optional[str],
) -> Optional[Dict[str, Any]]:
    """Convert a resolved cell dict into a canonical FMEA record.

    Returns ``None`` for rows that are entirely empty or have no
    ``component_type`` / ``failure_mode_name`` after stripping.
    """
    component_type = str(cells.get("component_type") or "").strip()
    fm_name = str(cells.get("failure_mode_name") or "").strip()

    mechanism = str(cells.get("failure_mechanism") or "").strip()

    if not component_type or not fm_name:
        return None  # blank or header-repeat row
    if not mechanism:
        raise ValueError(
            f"FMEA parse error in '{fmea_source_ref}' row {row_index}: "
            "required field 'failure_mechanism' is empty."
        )

    # Derive stable canonical ID.
    failure_mode_id = f"FM:{_slug(component_type)}:{_slug(fm_name)}"

    # Numeric risk fields.
    severity = _to_int(cells.get("severity"))
    occurrence = _to_int(cells.get("occurrence"))
    detection = _to_int(cells.get("detection"))
    rpn_raw = _to_int(cells.get("rpn"))
    # Derive RPN if the column is missing but all three components are present.
    rpn: Optional[int] = rpn_raw
    if rpn is None and all(x is not None for x in (severity, occurrence, detection)):
        rpn = (severity or 0) * (occurrence or 0) * (detection or 0)

    # Latency — stored in hours in the KG (hours is what _fetch_failure_modes returns).
    lat_min_days = _to_float(cells.get("expected_latency_min_days"))
    lat_max_days = _to_float(cells.get("expected_latency_max_days"))
    lat_min_hours = round(lat_min_days * 24, 2) if lat_min_days is not None else None
    lat_max_hours = round(lat_max_days * 24, 2) if lat_max_days is not None else None

    local_effect = str(cells.get("local_effect") or "").strip() or None
    system_effect = str(cells.get("system_effect") or "").strip() or None
    end_effect = str(cells.get("end_effect") or "").strip() or None
    expected_symptoms = _split_effect_to_symptoms(local_effect)
    anomaly_pattern = _resolve_anomaly_pattern(str(cells.get("expected_anomaly_pattern") or ""))

    return {
        "fmea_source_ref": fmea_source_ref,
        "component_type": component_type,
        "failure_mode_id": failure_mode_id,
        "failure_mode_name": fm_name,
        "item_function": str(cells.get("item_function") or "").strip() or None,
        "failure_mechanism": mechanism,
        "local_effect": local_effect,
        "system_effect": system_effect,
        "end_effect": end_effect,
        "potential_causes": _split_causes(cells.get("potential_causes")),
        "detection_method": str(cells.get("detection_method") or "").strip() or None,
        "severity": severity,
        "occurrence": occurrence,
        "detection_rating": _to_int(cells.get("detection_rating")) or detection,
        "detection": _to_int(cells.get("detection_rating")) or detection,
        "rpn": rpn,
        "safety_function_impact": str(cells.get("safety_function_impact") or "").strip() or None,
        "tech_spec_applicability": str(cells.get("tech_spec_applicability") or "").strip() or None,
        "failure_rate": _to_float(cells.get("failure_rate")),
        "failure_mode_ratio": _to_float(cells.get("failure_mode_ratio")),
        "mission_time_hours": _to_float(cells.get("mission_time_hours")),
        "criticality": str(cells.get("criticality") or "").strip() or None,
        "expected_latency_min_hours": lat_min_hours,
        "expected_latency_max_hours": lat_max_hours,
        "expected_anomaly_pattern": anomaly_pattern,
        "fmea_revision_date": str(cells.get("fmea_revision_date") or "").strip() or None,
        "expected_symptoms": expected_symptoms,
        "corrective_actions": _split_actions(cells.get("corrective_actions")),
        "notes": str(cells.get("notes") or "").strip() or None,
        "_sheet": sheet,
        "_row_index": row_index,
    }


# ---------------------------------------------------------------------------
# Format-specific row readers
# ---------------------------------------------------------------------------

def _rows_from_csv(path: Path) -> List[Tuple[Optional[str], List[List[str]]]]:
    """Read a CSV file and return ``[(None, rows)]`` where rows is a list of
    cell lists (all strings)."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as fh:
                sample = fh.read(4096)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(fh, dialect)
                rows = list(reader)
            return [(None, rows)]
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Cannot decode CSV file: {path}")


def _rows_from_xlsx(path: Path) -> List[Tuple[Optional[str], List[List[Any]]]]:
    """Read all sheets from an .xlsx workbook using openpyxl."""
    try:
        import openpyxl  # already in project deps
    except ImportError as exc:
        raise ImportError("openpyxl is required to parse .xlsx files") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: List[Tuple[Optional[str], List[List[Any]]]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        # Skip empty sheets.
        if any(any(c is not None for c in r) for r in rows):
            result.append((sheet_name, rows))
    wb.close()
    return result


def _rows_from_xls(path: Path) -> List[Tuple[Optional[str], List[List[Any]]]]:
    """Read all sheets from an .xls workbook using xlrd."""
    try:
        import xlrd  # already in project deps
    except ImportError as exc:
        raise ImportError("xlrd is required to parse .xls files") from exc

    wb = xlrd.open_workbook(str(path))
    result: List[Tuple[Optional[str], List[List[Any]]]] = []
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = [ws.row_values(r) for r in range(ws.nrows)]
        if rows:
            result.append((sheet_name, rows))
    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_fmea_file(
    path: str | Path,
    *,
    column_map_override: Optional[Dict[str, List[str]]] = None,
    sheet_filter: Optional[Sequence[str]] = None,
    profile_name: str = "auto",
    include_normalization_metadata: bool = True,
) -> List[Dict[str, Any]]:
    """Parse a FMEA spreadsheet into a list of canonical FMEA record dicts.

    Args:
        path: Path to a ``.csv``, ``.xlsx``, or ``.xls`` file.
        column_map_override: Optional dict that is **merged** into
            :data:`DEFAULT_COLUMN_MAP`.  Keys must be canonical field names;
            values are lists of additional regex patterns to try before the
            defaults.  Use this to add plant-specific column naming without
            replacing the default patterns.
        sheet_filter: For multi-sheet workbooks, only parse the sheets whose
            names are in this list.  Pass ``None`` (default) to parse all
            sheets.

    Returns:
        List of record dicts.  Each dict contains at minimum:
        ``fmea_source_ref``, ``component_type``, ``failure_mode_id``,
        ``failure_mode_name``.

    Raises:
        ValueError: If required columns (``component_type``,
            ``failure_mode_name``) cannot be resolved in a sheet, or if the
            file extension is not recognised.
        FileNotFoundError: If *path* does not exist.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"FMEA file not found: {path}")

    # Build merged column map (override > profile > defaults).
    merged_map = _build_column_map(
        profile_name=profile_name,
        column_map_override=column_map_override,
    )

    resolver = FmeaColumnResolver(merged_map)
    fmea_source_ref = path.name
    suffix = path.suffix.lower()

    if suffix == ".csv":
        sheets = _rows_from_csv(path)
    elif suffix == ".xlsx":
        sheets = _rows_from_xlsx(path)
    elif suffix == ".xls":
        sheets = _rows_from_xls(path)
    else:
        raise ValueError(
            f"Unsupported file extension '{suffix}'. "
            f"Expected .csv, .xlsx, or .xls."
        )

    records: List[Dict[str, Any]] = []

    for sheet_name, raw_rows in sheets:
        if sheet_filter is not None and sheet_name not in sheet_filter:
            LOGGER.debug("Skipping sheet '%s' (not in sheet_filter).", sheet_name)
            continue

        # Find the header row: the first row that resolves at least one required field.
        header_row_idx: Optional[int] = None
        resolved_cols: Dict[str, int] = {}
        for i, row in enumerate(raw_rows):
            headers = [str(c) if c is not None else "" for c in row]
            candidate = resolver.resolve(headers)
            if "component_type" in candidate or "failure_mode_name" in candidate:
                header_row_idx = i
                resolved_cols = candidate
                break

        source_label = f"{fmea_source_ref}[{sheet_name}]" if sheet_name else fmea_source_ref
        if header_row_idx is None:
            LOGGER.warning(
                "No recognisable header row found in '%s' — sheet skipped. "
                "Check column names against DEFAULT_COLUMN_MAP.",
                source_label,
            )
            continue

        try:
            resolver.validate_required(resolved_cols, source_label)
        except ValueError:
            LOGGER.error(
                "Required columns missing in '%s'. Resolved so far: %s",
                source_label, sorted(resolved_cols.keys()),
            )
            raise

        sheet_records = 0
        for row_offset, row in enumerate(raw_rows[header_row_idx + 1:], start=1):
            cells = {
                canonical: row[col_idx] if col_idx < len(row) else None
                for canonical, col_idx in resolved_cols.items()
            }
            rec = _build_record(cells, row_offset, fmea_source_ref, sheet_name)
            if rec is not None:
                records.append(rec)
                sheet_records += 1

        LOGGER.info(
            "Parsed %d FMEA records from '%s'.",
            sheet_records, source_label,
        )

    LOGGER.info(
        "Total FMEA records from '%s': %d across %d sheet(s).",
        fmea_source_ref, len(records), len(sheets),
    )
    normalized, report = normalize_fmea_records(records, profile_name=profile_name)
    LOGGER.info(
        "FMEA normalization (%s): derived=%d, nlp_inferred=%d, critical_missing=%d",
        report.get("profile_used"),
        int(report.get("derived_field_count", 0) or 0),
        int(report.get("nlp_inferred_field_count", 0) or 0),
        int(report.get("critical_field_missing_count", 0) or 0),
    )
    if include_normalization_metadata:
        for rec in normalized:
            rec["_fmea_ingestion_quality"] = report
    else:
        for rec in normalized:
            rec.pop("_field_quality", None)
            rec.pop("_normalization_profile", None)
    return normalized


def parse_fmea_files(
    paths: Sequence[str | Path],
    **kwargs: Any,
) -> List[Dict[str, Any]]:
    """Parse multiple FMEA files and return a combined record list.

    Args:
        paths: Iterable of file paths.
        **kwargs: Forwarded to :func:`parse_fmea_file`.

    Returns:
        Combined list of all records from all files.
    """
    all_records: List[Dict[str, Any]] = []
    merged_report: Optional[Dict[str, Any]] = None
    for p in paths:
        records = parse_fmea_file(p, **kwargs)
        all_records.extend(records)
        file_report: Optional[Dict[str, Any]] = None
        if records:
            maybe = records[0].get("_fmea_ingestion_quality")
            if isinstance(maybe, dict):
                file_report = maybe
        if file_report:
            merged_report = _merge_ingestion_reports(merged_report, file_report)
    if merged_report:
        for rec in all_records:
            rec["_fmea_ingestion_quality"] = merged_report
    return all_records


def _merge_ingestion_reports(
    current: Optional[Dict[str, Any]],
    incoming: Dict[str, Any],
) -> Dict[str, Any]:
    if current is None:
        return dict(incoming)

    merged = dict(current)
    merged["total_fms_ingested"] = int(current.get("total_fms_ingested", 0) or 0) + int(
        incoming.get("total_fms_ingested", 0) or 0
    )
    merged["critical_field_missing_count"] = int(
        current.get("critical_field_missing_count", 0) or 0
    ) + int(incoming.get("critical_field_missing_count", 0) or 0)
    merged["enrichment_field_missing_count"] = int(
        current.get("enrichment_field_missing_count", 0) or 0
    ) + int(incoming.get("enrichment_field_missing_count", 0) or 0)
    merged["derived_field_count"] = int(current.get("derived_field_count", 0) or 0) + int(
        incoming.get("derived_field_count", 0) or 0
    )
    merged["nlp_inferred_field_count"] = int(
        current.get("nlp_inferred_field_count", 0) or 0
    ) + int(incoming.get("nlp_inferred_field_count", 0) or 0)
    merged["orphaned_fm_count"] = int(current.get("orphaned_fm_count", 0) or 0) + int(
        incoming.get("orphaned_fm_count", 0) or 0
    )

    cur_profile = str(current.get("profile_used") or "auto")
    in_profile = str(incoming.get("profile_used") or "auto")
    merged["profile_used"] = cur_profile if cur_profile == in_profile else "mixed"

    total = merged["total_fms_ingested"]
    cur_weight = int(current.get("total_fms_ingested", 0) or 0)
    in_weight = int(incoming.get("total_fms_ingested", 0) or 0)
    cur_conf = float(current.get("format_autodetect_confidence", 0.0) or 0.0)
    in_conf = float(incoming.get("format_autodetect_confidence", 0.0) or 0.0)
    if total > 0:
        merged["format_autodetect_confidence"] = ((cur_conf * cur_weight) + (in_conf * in_weight)) / total
    else:
        merged["format_autodetect_confidence"] = 0.0
    return merged
